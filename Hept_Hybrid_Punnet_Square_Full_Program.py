print("[INFO] Command loading")
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
fontStyle = Font(size = "14")
time.sleep(1)
print("[INFO] Command loaded")
print("")
print("[INFO] Creating the Excel Sheet")
wb = Workbook()
sheet = wb.active
time.sleep(2)
print("[INFO] Excel sheet created")
print("")
print("[INFO] Initializing Allele creation.")
n = 7
r = 2
for i in range(0, 2**n):
    output = f'{i:07b}'.replace('0', 'H').replace('1', 'T')
    sheet.cell(row=r, column=2).value = output
    r=r+1

wb.save("hept_hybrid.xlsx")
#Value exchange code
#Alleles = TPYRIGVHIJKLM(X,Y)
r=2
c=2
for x in range(128):
      row = sheet.cell(row=r, column=2).value
      sheet.cell(row=r, column=2).value = ''


      a1=row[0]
      a2=row[1]
      a3=row[2]
      a4=row[3]
      a5=row[4]
      a6=row[5]
      a7=row[6]

      #Val a1
      if a1 == "H":
            a1 = "T"
      else:
            a1 = "t"
      #Val a2
      if a2 == "H":
            a2 = "P"
      else:
            a2 = "p"
      #Val a3
      if a3 == "H":
            a3 = "Y"
      else:
            a3 = "y"
      #Val a4
      if a4 == "H":
            a4 = "R"
      else:
            a4 = "r"
      #Val a5
      if a5 == "H":
            a5 = "I"
      else:
            a5 = "i"
      #Val a6
      if a6 == "H":
            a6 = "G"
      else:
            a6 = "g"
      #Val a7
      if a7 == "H":
            a7 = "V"
      else:
            a7 = "v"

      #output
      output = a1+a2+a3+a4+a5+a6+a7
      sheet.cell(row=r, column=1).value = output
      sheet.cell(row=1, column=c).value = output
      #print (output)
      r=r+1
      c=c+1
wb.save("hept_hybrid.xlsx")
print("[PROCESS] Allele Creation Completed\n")
print("[PROCESS] Starting module for calculation of offsprings.\n")
time.sleep(2)
print("[PROCESS] Load Complete.\n")
print("[PROCESS] Process - Form all the offsprings.py NOW RUNNING\n")
time.sleep(3)
print("[PROCESS] Loading resources into CPU...")
time.sleep(2)
print("[ERROR] PROCESSOR MAXED OUT.\n")
time.sleep(2)
print("[BACKUP] OVERCLOCKING CPU - PERFORMING FINAL OPERATION\n")
time.sleep(1)
r=1
c=2
r1=2
counter=0
for l in range(16384):

      row = sheet.cell(row=r, column=c).value
      col = sheet.cell(row=r1, column=1).value

      a1=row[0]
      a2=row[1]
      a3=row[2]
      a4=row[3]
      a5=row[4]
      a6=row[5]
      a7=row[6]

      A1=col[0]
      A2=col[1]
      A3=col[2]
      A4=col[3]
      A5=col[4]
      A6=col[5]
      A7=col[6]


      output = a1+A1+a2+A2+a3+A3+a4+A4+a5+A5+a6+A6+a7+A7
      sheet.cell(row=r1, column=(2+counter)).value = output
      if r1 == 129:
            r1=1
            c=c+1
            counter=counter+1
            print("No.of rows done : ",counter)
            #time.sleep(1)
      r1=r1+1
wb.save("hept_hybrid.xlsx")
print("")
print("[PROCESS] The program has completed\n")
print("Copyright (c) Meit Sant 2021")
time.sleep(30)